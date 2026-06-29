import os
import asyncio
import uuid
from pathlib import Path
from typing import Optional

import json

from app.config import settings
from app.logger import get_logger

logger = get_logger(__name__)


def _parse_docx_text(file_path: str) -> str:
    """Extract text from .docx using LightRAG native parser (python-docx)."""
    from lightrag.parser.docx.parse_document import extract_docx_blocks

    blocks = extract_docx_blocks(file_path)
    lines: list[str] = []
    for b in blocks:
        heading = b.get("heading", "")
        content = b.get("content", "")
        if heading:
            lines.append(f"## {heading}")
        if content:
            lines.append(content)
    return "\n\n".join(lines)


def _convert_doc_to_docx(doc_path: str) -> str:
    """Convert .doc to .docx via LibreOffice headless, return .docx path."""
    import subprocess

    out_dir = os.path.dirname(doc_path) or "."
    result = subprocess.run(
        [
            "libreoffice", "--headless", "--convert-to", "docx",
            "--outdir", out_dir, doc_path,
        ],
        capture_output=True, text=True, timeout=60,
    )
    if result.returncode != 0:
        raise RuntimeError(f"LibreOffice conversion failed: {result.stderr}")
    base = os.path.splitext(doc_path)[0]
    docx_path = base + ".docx"
    if not os.path.exists(docx_path):
        raise RuntimeError(f"Converted .docx not found at {docx_path}")
    return docx_path


def _create_llm_model_func():
    from openai import OpenAI as SyncOpenAI

    client = SyncOpenAI(
        api_key=settings.openai_api_key,
        base_url=settings.openai_base_url,
    )

    async def llm_func(prompt: str, system_prompt: str | None = None, **kwargs) -> str:
        messages = []
        if system_prompt:
            messages.append({"role": "system", "content": system_prompt})
        messages.append({"role": "user", "content": prompt})

        model = settings.model_name or "qwen3.6-35B-A3B-apex"
        response = client.chat.completions.create(
            model=model,
            messages=messages,
            temperature=kwargs.get("temperature", 0.1),
            max_tokens=kwargs.get("max_tokens", 4096),
        )
        return response.choices[0].message.content or ""

    return llm_func


def _create_embedding_func():
    import asyncio
    import numpy as np
    import httpx
    from lightrag.utils import wrap_embedding_func_with_attrs
    base_url = settings.embedding_base_url.rstrip("/").removesuffix("/v1")
    api_key = settings.embedding_api_key

    embedding_dim_val = settings.embedding_dim
    embedding_model_val = settings.embedding_model

    @wrap_embedding_func_with_attrs(embedding_dim=embedding_dim_val, max_token_size=8192)
    async def emb_func(texts: list[str]) -> np.ndarray:
        if isinstance(texts, str):
            texts = [texts]
        data = {"model": embedding_model_val, "input": texts}
        headers = {"Authorization": f"Bearer {api_key}"} if api_key else {}
        url = f"{base_url}/v1/embeddings"
        max_retries = 3
        last_error = None
        for attempt in range(max_retries):
            try:
                async with httpx.AsyncClient(timeout=120.0) as client:
                    resp = await client.post(url, json=data, headers=headers)
                    if resp.status_code == 200:
                        break
                    last_error = RuntimeError(
                        f"Embedding API error {resp.status_code}: {resp.text[:300]}"
                    )
                    if resp.status_code < 500:
                        raise last_error
            except httpx.TimeoutException as e:
                last_error = RuntimeError(f"Embedding API timeout: {url}")
            except RuntimeError:
                raise
            except Exception as e:
                last_error = e
            if attempt < max_retries - 1:
                wait = 2 ** attempt
                logger.warning(
                    f"[RAG:Embed] Attempt {attempt + 1}/{max_retries} failed, "
                    f"retrying in {wait}s..."
                )
                await asyncio.sleep(wait)
        else:
            raise last_error  # type: ignore[misc]

        body = resp.json()
        sorted_data = sorted(body["data"], key=lambda x: x["index"])
        embeddings = [d["embedding"] for d in sorted_data]
        return np.array(embeddings, dtype=np.float32)

    return emb_func


async def _fast_insert_text(lightrag, text: str, doc_id: str, file_path: str) -> int:
    """Insert text into LightRAG with chunking + embedding only, no entity extraction.

    Much faster than the full pipeline (~50-90 % less LLM calls).
    Retrieval works with LightRAG naive / mix query modes.
    """
    from lightrag.utils import compute_mdhash_id, sanitize_text_for_encoding

    # normalize_document_file_path moved between LightRAG versions
    try:
        from lightrag.parser.routing import normalize_document_file_path
    except ImportError:
        from lightrag.utils_pipeline import normalize_document_file_path

    emb_func = _create_embedding_func()
    text = sanitize_text_for_encoding(text)
    file_path = normalize_document_file_path(file_path)

    # Simple paragraph-based chunking with token-aware split
    chunk_token_size = 800
    paragraphs = text.split("\n\n")
    chunks: list[str] = []
    current = ""
    current_tokens = 0
    for para in paragraphs:
        para_tokens = len(lightrag.tokenizer.encode(para))
        if current_tokens + para_tokens > chunk_token_size and current:
            chunks.append(current.strip())
            current = para
            current_tokens = para_tokens
        else:
            current = current + "\n\n" + para if current else para
            current_tokens += para_tokens
    if current.strip():
        chunks.append(current.strip())

    if not chunks:
        logger.warning(f"[RAG:FastInsert] No chunks produced for {file_path}")
        return 0

    # Generate doc_id
    doc_key = compute_mdhash_id(text, prefix="doc-") if not doc_id else doc_id

    # Deduplicate: filter_keys returns keys NOT in storage (need insertion).
    # If doc_key is NOT in the returned set, it already exists → skip.
    new_keys = await lightrag.full_docs.filter_keys({doc_key})
    if doc_key not in new_keys:
        logger.info(f"[RAG:FastInsert] Doc {doc_key} already in storage, skip")
        return 0

    # Build chunk entries
    inserting_chunks: dict[str, dict] = {}
    for idx, chunk_text in enumerate(chunks):
        chunk_key = compute_mdhash_id(chunk_text, prefix="chunk-")
        tokens = len(lightrag.tokenizer.encode(chunk_text))
        inserting_chunks[chunk_key] = {
            "content": chunk_text,
            "full_doc_id": doc_key,
            "tokens": tokens,
            "chunk_order_index": idx,
            "file_path": file_path,
        }

    # Generate embeddings for all chunks at once
    chunk_texts = [c["content"] for c in inserting_chunks.values()]
    embeddings = await emb_func(chunk_texts)
    for key, emb in zip(inserting_chunks.keys(), embeddings):
        inserting_chunks[key]["vector"] = emb.tolist()

    # Write to LightRAG storages
    new_docs = {doc_key: {"content": text, "file_path": file_path}}
    await asyncio.gather(
        lightrag.chunks_vdb.upsert(inserting_chunks),
        lightrag.text_chunks.upsert(inserting_chunks),
        lightrag.full_docs.upsert(new_docs),
    )

    # Flush to disk — critical for multi-worker and persistence across restarts
    await asyncio.gather(
        lightrag.chunks_vdb.index_done_callback(),
        lightrag.text_chunks.index_done_callback(),
        lightrag.full_docs.index_done_callback(),
    )

    logger.info(
        f"[RAG:FastInsert] Indexed {len(chunks)} chunks "
        f"({sum(c['tokens'] for c in inserting_chunks.values())} tokens)"
    )
    return len(chunks)


class RAGEngineAdapter:

    def __init__(self):
        self._rags: dict[str, object] = {}

    async def _init_user_rag(self, user_id: str):
        if user_id in self._rags:
            return

        from raganything import RAGAnything, RAGAnythingConfig

        base_dir = os.path.join(settings.upload_dir, ".rag_storage")
        working_dir = os.path.join(base_dir, user_id)
        parser_output_dir = os.path.join(settings.upload_dir, ".rag_parse_output", user_id)
        os.makedirs(working_dir, exist_ok=True)
        os.makedirs(parser_output_dir, exist_ok=True)

        config = RAGAnythingConfig(
            working_dir=working_dir,
            parser_output_dir=parser_output_dir,
            parser="mineru",
            parse_method="ocr",
            enable_image_processing=False,
            enable_table_processing=False,
            enable_equation_processing=False,
            max_concurrent_files=1,
            max_context_tokens=settings.rag_max_context_tokens,
            use_full_path=True,
        )

        rag = RAGAnything(
            config=config,
            llm_model_func=_create_llm_model_func(),
            embedding_func=_create_embedding_func(),
            lightrag_kwargs={
                "workspace": user_id,
                "embedding_func_max_async": settings.embedding_workers,
                "embedding_batch_num": 16,
                "chunk_top_k": settings.rag_chunk_top_k,
                "cosine_threshold": settings.rag_cosine_threshold,
                "max_parallel_insert": 1,
                "enable_llm_cache": True,
                "enable_llm_cache_for_entity_extract": True,
            },
        )

        await rag._ensure_lightrag_initialized()
        self._rags[user_id] = rag

    async def _ensure_ready(self, user_id: str):
        if user_id not in self._rags:
            try:
                await self._init_user_rag(user_id)
            except Exception as e:
                logger.error(f"[RAG:Init] Init failed for {user_id}: {e}")

    def _build_category_map(self) -> dict[str, str]:
        """Build file_path → category mapping from .meta files on disk."""
        mapping: dict[str, str] = {}
        upload_dir = Path(settings.upload_dir)
        if not upload_dir.exists():
            return mapping
        for meta_path in upload_dir.glob("*.meta"):
            try:
                with open(meta_path, "r", encoding="utf-8") as f:
                    meta = json.load(f)
            except Exception:
                continue
            category = meta.get("category", "")
            if category == "上传文件":
                category = ""
            original_name = meta.get("original_name", "")
            if not original_name:
                continue
            file_id = meta_path.stem
            for fpath in upload_dir.iterdir():
                if not fpath.is_file() or fpath.suffix == ".meta":
                    continue
                if fpath.name.startswith(f"{file_id}_"):
                    mapping[str(fpath.resolve())] = category
                    break
        return mapping

    async def index_file(
        self, file_path: str, file_name: str | None = None,
        category: str = "", user_id: str = "default"
    ) -> tuple[str, int]:
        await self._ensure_ready(user_id)
        user_rag = self._rags[user_id]

        import hashlib
        doc_id = hashlib.md5(str(file_path).encode()).hexdigest()[:16]
        ext = Path(file_path).suffix.lower()

        # Bypass the slow MinerU PDF pipeline for text/Excel: extract content and
        # insert directly into LightRAG.
        text_content = None
        if ext in (".txt", ".md", ".csv"):
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    text_content = f.read()
                if not text_content.strip():
                    logger.warning(f"[RAG:Index] Empty file: {file_path}")
                    return "", 0
            except Exception as e:
                logger.error(f"[RAG:Index] Read error for {file_path}: {e}")
                return "", 0

        elif ext in (".xlsx", ".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, data_only=True)
                lines = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    lines.append(f"=== Sheet: {sheet_name} ===")
                    for row in ws.iter_rows(values_only=True):
                        row_str = " | ".join(str(cell) if cell is not None else "" for cell in row)
                        if row_str.strip():
                            lines.append(row_str)
                text_content = "\n".join(lines)
                wb.close()
                if not text_content.strip():
                    logger.warning(f"[RAG:Index] Empty Excel file: {file_path}")
                    return "", 0
                logger.info(
                    f"[RAG:Index] Extracted Excel {file_name or Path(file_path).name}: "
                    f"{len(lines)} lines, {len(text_content)} chars"
                )
            except Exception as e:
                logger.error(f"[RAG:Index] Excel extraction error for {file_path}: {e}")
                return "", 0

        elif ext in (".docx",):
            try:
                text_content = _parse_docx_text(file_path)
                if not text_content.strip():
                    logger.warning(f"[RAG:Index] Empty docx: {file_path}")
                    return "", 0
                logger.info(
                    f"[RAG:Index] Native-parsed DOCX {file_name or Path(file_path).name}: "
                    f"{len(text_content)} chars"
                )
            except Exception as e:
                logger.error(f"[RAG:Index] DOCX native parse error for {file_path}: {e}")
                return "", 0

        elif ext in (".doc",):
            try:
                docx_path = _convert_doc_to_docx(file_path)
                text_content = _parse_docx_text(docx_path)
                os.remove(file_path)
                file_path = docx_path
                if not text_content.strip():
                    logger.warning(f"[RAG:Index] Empty doc (converted): {file_path}")
                    return "", 0
                logger.info(
                    f"[RAG:Index] Converted DOC→DOCX {file_name or Path(file_path).name}: "
                    f"{len(text_content)} chars"
                )
            except Exception as e:
                logger.error(f"[RAG:Index] DOC conversion error for {file_path}: {e}")
                return "", 0

        chunk_count = 0
        if text_content is not None:
            if settings.rag_fast_indexing:
                # Fast: chunk + embed only, skip entity extraction
                try:
                    chunk_count = await _fast_insert_text(
                        lightrag=user_rag.lightrag,
                        text=text_content,
                        doc_id=doc_id,
                        file_path=file_path,
                    )
                    logger.info(
                        f"[RAG:FastIndex] Fast-indexed {file_name or Path(file_path).name} "
                        f"({len(text_content)} chars, {chunk_count} chunks)"
                    )
                except Exception as e:
                    logger.error(f"[RAG:FastIndex] Fast insert error for {file_name or Path(file_path).name}: {e}")
                    raise RuntimeError(f"快速索引写入失败: {e}") from e
            else:
                # Full pipeline: chunking → entity extraction → relation graph → vector index
                from raganything.utils import insert_text_content
                try:
                    split_char = '\n\n' if ext in ('.txt', '.md') else None
                    await insert_text_content(
                        lightrag=user_rag.lightrag,
                        input=text_content,
                        ids=doc_id,
                        file_paths=file_path,
                        split_by_character=split_char,
                    )
                except Exception as e:
                    logger.error(f"[RAG:Index] insert_text_content error for {file_name or Path(file_path).name}: {e}")
                    raise RuntimeError(f"文本索引写入失败: {e}") from e

                # Estimate chunk count (LightRAG ~1200 tokens ≈ 500-600 chars per chunk)
                chunk_count = max(1, len(text_content) // 500)
                logger.info(
                    f"[RAG:Index] Full-pipeline indexed {file_name or Path(file_path).name} "
                    f"({len(text_content)} chars, ~{chunk_count} chunks)"
                )
        else:
            try:
                await user_rag.process_document_complete(
                    file_path=file_path,
                    file_name=file_name or Path(file_path).name,
                    doc_id=f"{category}::{doc_id}",
                )
            except Exception as e:
                logger.error(f"[RAG:Index] Error indexing (MinerU) {file_path}: {e}")
                return "", 0
            # Rough estimate for MinerU-parsed documents
            try:
                file_size = os.path.getsize(file_path)
                chunk_count = max(1, file_size // 500)
            except Exception:
                chunk_count = 1

        # Store mapping in-memory (keyed by user_id + file_name)
        _doc_id_map[f"{user_id}:{file_name or Path(file_path).name}"] = f"{category}::{doc_id}"
        return doc_id, chunk_count

    async def search(
        self, query_text: str, category: str | None = None, top_k: int = 3,
        user_id: str = "default"
    ) -> list[dict]:
        await self._ensure_ready(user_id)
        lightrag = self._rags[user_id].lightrag

        try:
            from lightrag.base import QueryParam

            if category:
                return await self._search_with_category(
                    lightrag, query_text, category, top_k, user_id
                )

            mode = "naive" if settings.rag_fast_indexing else "mix"
            param = QueryParam(
                mode=mode,
                only_need_context=False,
                top_k=settings.rag_chunk_top_k,
                chunk_top_k=settings.rag_chunk_top_k,
                enable_rerank=False,
            )
            result_text = await asyncio.wait_for(
                lightrag.aquery(query_text, param=param),
                timeout=settings.rag_query_timeout,
            )
        except asyncio.TimeoutError:
            logger.warning(f"[RAG:Query] Timeout after {settings.rag_query_timeout}s")
            return []
        except Exception as e:
            logger.error(f"[RAG:Query] Error: {e}")
            return []

        if not result_text or not result_text.strip():
            return []

        import json
        import re

        content_parts = []

        for block in re.findall(r"```(?:json)?\s*({.*?})\s*```", result_text, re.DOTALL):
            try:
                obj = json.loads(block)
                content = obj.get("content", "")
                if content:
                    content_parts.append(content)
            except json.JSONDecodeError:
                pass
        if not content_parts:
            content_parts.append(result_text.strip())

        seen = set()
        results = []
        for text in content_parts:
            text = text.strip()
            if not text or text in seen:
                continue
            seen.add(text)
            results.append({
                "text": text,
                "file_name": "知识库",
                "category": category or "全部",
                "score": 1.0,
            })
            if len(results) >= top_k:
                break

        return results

    async def _search_with_category(
        self, lightrag, query_text: str, category: str, top_k: int, user_id: str
    ) -> list[dict]:
        """Search LightRAG with structured results, then filter by category."""
        from lightrag.base import QueryParam

        mode = "naive" if settings.rag_fast_indexing else "mix"
        param = QueryParam(
            mode=mode,
            only_need_context=False,
            top_k=settings.rag_chunk_top_k,
            chunk_top_k=settings.rag_chunk_top_k,
            enable_rerank=False,
        )

        try:
            data_result = await asyncio.wait_for(
                lightrag.aquery_data(query_text, param=param),
                timeout=settings.rag_query_timeout,
            )
        except asyncio.TimeoutError:
            logger.warning(f"[RAG:Query] aquery_data timeout after {settings.rag_query_timeout}s")
            return []
        except Exception as e:
            logger.error(f"[RAG:Query] aquery_data error: {e}")
            return []

        chunks = []
        if data_result.get("status") == "success":
            chunks = data_result.get("data", {}).get("chunks", [])

        if not chunks:
            return []

        cat_map = self._build_category_map()
        # LightRAG's normalize_document_file_path stores basename, but
        # cat_map keys are absolute paths. Build basename fallback map.
        basename_map: dict[str, str] = {}
        for fp_abs, fp_cat in cat_map.items():
            bn = Path(fp_abs).name
            if bn not in basename_map:
                basename_map[bn] = fp_cat

        matched = []
        for c in chunks:
            fp = c.get("file_path", "")
            chunk_cat = cat_map.get(fp, "")
            if not chunk_cat:
                chunk_cat = basename_map.get(Path(fp).name, "")
            if chunk_cat == category:
                matched.append(c)

        if not matched:
            return []

        seen = set()
        results = []
        for c in matched:
            text = c.get("content", "").strip()
            if not text or text in seen:
                continue
            seen.add(text)
            fp = c.get("file_path", "")
            file_name = Path(fp).name if fp else "知识库"
            results.append({
                "text": text,
                "file_name": file_name,
                "category": category,
                "score": 1.0,
            })
            if len(results) >= top_k:
                break

        return results

    async def search_text(
        self, query_text: str, category: str | None = None, top_k: int = 5,
        user_id: str = "default"
    ) -> list[dict]:
        return await self.search(query_text, category=category, top_k=top_k, user_id=user_id)

    async def stats(self, user_id: str = "default") -> dict:
        await self._ensure_ready(user_id)
        return self._rags[user_id].get_config_info() if user_id in self._rags else {}

    async def remove_file(self, doc_id: str, user_id: str = "default") -> bool:
        await self._ensure_ready(user_id)
        lightrag = self._rags[user_id].lightrag
        try:
            result = await lightrag.adelete_by_doc_id(doc_id)
            if result.status == "success":
                logger.info(f"[RAG:Delete] Deleted doc {doc_id}: {result.message}")
                return True
            elif result.status == "not_found":
                logger.info(f"[RAG:Delete] Doc {doc_id} not found in RAG storage, treating as deleted")
                return True
            else:
                logger.error(f"[RAG:Delete] Deletion failed for {doc_id}: {result.message}")
                return False
        except Exception as e:
            logger.error(f"[RAG:Delete] Error deleting {doc_id}: {e}")
            return False


rag = RAGEngineAdapter()

_doc_id_map: dict[str, str] = {}
