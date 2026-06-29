from __future__ import annotations
import json
import asyncio

from sqlalchemy.orm import Session
from app.database import SessionLocal
from app.config import settings
from app.rag_engine import rag
from app.models.db_connection import DbConnection
from app.services import db_service
from app.services import experience_service
from app.logger import get_logger
from pathlib import Path
import os

logger = get_logger(__name__)


async def query_rag(query_text: str, category: str = "", top_k: int = 5,
                    user_id: str = "default", kb_scope: str = "none") -> str:
    """Search uploaded documents.

    Search order:
      1. If kb_scope="public": Public KB → filtered by category (exact tag match)
      2. Personal KB → always searched WITHOUT category filter (all files)

    Personal KB is always accessible to every user. Public KB only when
    the admin grants kb_scope="public".
    """
    include_public = (kb_scope == "public")
    category_param = category if category else None

    async def _do_search(effective_user: str, cat: str | None = None) -> list[dict]:
        results = await rag.search(query_text, category=cat,
                                   top_k=top_k, user_id=effective_user)
        return list(results) if results else []

    try:
        all_results: list[dict] = []

        if include_public and user_id != "default":
            # 1. Public KB: search with category filter (exact tag match)
            pub_results = await asyncio.wait_for(
                _do_search("default", cat=category_param),
                timeout=settings.rag_query_timeout + 5,
            )
            all_results.extend(pub_results)

        # 2. Personal KB: always search WITHOUT category filter (all files)
        pers_results = await asyncio.wait_for(
            _do_search(user_id, cat=None),
            timeout=settings.rag_query_timeout + 5,
        )
        seen = {r.get("text", "") for r in all_results}
        for r in pers_results:
            if r.get("text", "") not in seen:
                all_results.append(r)

        # 3. Fallback: if category filter yielded nothing in public KB,
        #    retry public KB without filter
        if not all_results and category_param:
            logger.warning(
                f"[RAG:Tool] Category '{category}' matched nothing in public KB, "
                f"retrying public KB without category filter"
            )
            pub_retry = await asyncio.wait_for(
                _do_search("default", cat=None),
                timeout=settings.rag_query_timeout + 5,
            )
            seen = {r.get("text", "") for r in all_results}
            for r in pub_retry:
                if r.get("text", "") not in seen:
                    all_results.append(r)

        if not all_results:
            if category:
                return (
                    f"No content found with tag '{category}'. "
                    f"Also checked all untagged documents — no relevant results. "
                    f"Try rephrasing the question."
                )
            return "No relevant document content found in the uploaded files."

        formatted_parts = [f"**Document Search Results ({len(all_results)} chunks):**\n"]
        for i, chunk in enumerate(all_results[:top_k], 1):
            text = chunk.get("text", "")
            source = chunk.get("file_name", "Unknown file")
            cat = chunk.get("category", "")
            score = chunk.get("score", 0)
            formatted_parts.append(
                f"[{i}] **Source:** {source} (category: {cat}, relevance: {score:.3f})\n    {text[:500]}"
            )
        return "\n\n".join(formatted_parts)

    except asyncio.TimeoutError:
        return "知识库检索超时，可能嵌入服务未启动，请稍后重试或联系管理员。"
    except Exception as e:
        return f"[RAG Engine Error] {str(e)}"


def _check_db_access(connection_id: int, db_scope: list[int] | None) -> bool:
    if db_scope is None:
        return True
    return connection_id in db_scope


async def list_db_connections(db_scope: list[int] | None = None) -> str:
    if db_scope is not None and len(db_scope) == 0:
        return ("数据库查询已被管理员限制，您当前无权使用此功能。"
                "请联系管理员开通权限。")

    db: Session = SessionLocal()
    try:
        conns = db.query(DbConnection).filter(
            DbConnection.status == "connected"
        ).all()

        if not conns:
            return ("No connected databases found. "
                    "Ask the admin to add and connect databases in '数据库管理'.")

        visible = conns
        if db_scope is not None:
            visible = [c for c in conns if c.id in db_scope]

        if not visible:
            return ("No database connections available within your access scope. "
                    "Contact the admin to adjust permissions.")

        lines = ["**Available Database Connections:**\n"]
        for c in visible:
            env_label = "🔬测试" if c.environment == "test" else "🏭生产"
            fields_count = 0
            if c.table_fields:
                try:
                    fields_count = len(json.loads(c.table_fields))
                except Exception:
                    pass
            lines.append(
                f"- ID:{c.id} | {c.name} | {env_label} | 表: `{c.table_name}` "
                f"({fields_count} 字段) | {c.host}:{c.port}"
            )
        return "\n".join(lines)
    except Exception as e:
        return f"[Database Error] Could not list connections: {str(e)}"
    finally:
        db.close()


async def list_db_tables(connection_id: int,
                         db_scope: list[int] | None = None) -> str:
    if db_scope is not None and len(db_scope) == 0:
        return ("数据库查询已被管理员限制，您当前无权使用此功能。"
                "请联系管理员开通权限。")

    if not _check_db_access(connection_id, db_scope):
        return (f"Access denied: connection {connection_id} is not within "
                "your authorised database scope. "
                "Please use list_db_connections to see available databases.")

    db: Session = SessionLocal()
    try:
        conn = db.query(DbConnection).filter(DbConnection.id == connection_id).first()
        if not conn:
            return f"Connection ID {connection_id} not found."

        if conn.status != "connected":
            return f"Connection '{conn.name}' is not connected. Status: {conn.status}"

        result = db_service.list_tables(
            host=conn.host, port=conn.port,
            user=conn.db_user, password=conn.db_password,
        )

        if not result["success"]:
            return f"Could not list tables: {result.get('message', 'Unknown error')}"

        tables = result["tables"]
        if not tables:
            return f"No tables found in database '{conn.name}' ({conn.host}:{conn.port})."

        lines = [f"**Tables in '{conn.name}' ({conn.host}:{conn.port}):**\n"]
        for t in tables:
            lines.append(f"- `{t['database']}`.`{t['table']}`")

        if conn.table_fields:
            try:
                fields = json.loads(conn.table_fields)
                field_info = ", ".join(f"{f['name']} ({f['type']})" for f in fields)
                lines.append(f"\n**Table `{conn.table_name}` fields:** {field_info}")
            except Exception:
                pass

        return "\n".join(lines)
    except Exception as e:
        return f"[Database Error] {str(e)}"
    finally:
        db.close()


async def query_db(sql_query: str, connection_id: int,
                   db_scope: list[int] | None = None) -> str:
    if db_scope is not None and len(db_scope) == 0:
        return ("数据库查询已被管理员限制，您当前无权使用此功能。"
                "请联系管理员开通权限。")

    if not _check_db_access(connection_id, db_scope):
        return (f"Access denied: connection {connection_id} is not within "
                "your authorised database scope. "
                "Please use list_db_connections to see available databases.")

    db: Session = SessionLocal()
    try:
        conn = db.query(DbConnection).filter(DbConnection.id == connection_id).first()
        if not conn:
            return f"Connection ID {connection_id} not found."

        if conn.status != "connected":
            return f"Connection '{conn.name}' is not connected."

        result = db_service.execute_query(
            host=conn.host, port=conn.port,
            user=conn.db_user, password=conn.db_password,
            table_name=conn.table_name,
            sql_query=sql_query,
        )

        if not result["success"]:
            return f"Query failed: {result.get('message', 'Unknown error')}"

        data = result["data"]
        columns = result["columns"]

        if not data:
            return "Query executed successfully but returned no rows."

        header = "| " + " | ".join(columns) + " |\n"
        separator = "| " + " | ".join(["---"] * len(columns)) + " |\n"
        body_lines = []
        for row in data:
            body_lines.append("| " + " | ".join(
                str(row.get(c, "")) if row.get(c) is not None else "" for c in columns
            ) + " |")

        result_text = (
            f"**Database Query Results ({len(data)} rows):**\n\n"
            f"{header}{separator}"
            + "\n".join(body_lines[:50])
        )
        if len(data) > 50:
            result_text += f"\n\n*(Showing first 50 of {len(data)} rows)*"
        return result_text

    except ValueError as e:
        return f"[SQL Validation Error] {str(e)}"
    except Exception as e:
        return f"[Database Error] {str(e)}"
    finally:
        db.close()


async def find_file_by_name(keyword: str, user_id: str = "default") -> str:
    """Search uploaded files by name keyword, returning file content directly.

    Unlike RAG-based search, this tool bypasses vector search and reads
    the actual file contents. Use this when the prompt template explicitly
    specifies a particular file to look up (e.g., "if the user mentions
    turnover, look for 离职报告.pdf").

    The keyword is matched against file names case-insensitively as a
    substring search. The first matching file's content is returned.
    """
    upload_dir = Path(settings.upload_dir)
    if not upload_dir.exists():
        return "未找到上传目录。"

    matches = []
    kw = keyword.lower().strip()
    for fpath in upload_dir.iterdir():
        if not fpath.is_file() or fpath.suffix == ".meta" or fpath.name.startswith("."):
            continue
        if kw in fpath.name.lower():
            original_name = fpath.name
            meta_path = upload_dir / f"{fpath.stem.rsplit('_', 1)[0] if '_' in fpath.stem else fpath.stem}.meta"
            if meta_path.exists():
                try:
                    with open(meta_path, "r", encoding="utf-8") as f:
                        meta = json.load(f)
                    original_name = meta.get("original_name", fpath.name)
                except Exception:
                    pass
            matches.append((fpath, original_name))

    if not matches:
        return f"未找到名称包含「{keyword}」的文件。"

    fpath, original_name = matches[0]

    try:
        ext = fpath.suffix.lower()
        if ext in (".txt", ".md", ".csv"):
            with open(fpath, "r", encoding="utf-8") as f:
                content = f.read()
        elif ext in (".xlsx", ".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(str(fpath), data_only=True)
            lines = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                lines.append(f"=== Sheet: {sheet_name} ===")
                for row in ws.iter_rows(values_only=True):
                    row_str = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    if row_str.strip():
                        lines.append(row_str)
            content = "\n".join(lines)
            wb.close()
        elif ext == ".pdf":
            try:
                import pymupdf
                doc = pymupdf.open(str(fpath))
                pages = [page.get_text() for page in doc]
                content = "\n\n".join(pages)
                doc.close()
            except ImportError:
                return f"PDF 文件需安装 pymupdf 库才能读取：{original_name}"
        elif ext in (".docx", ".doc"):
            try:
                import docx
                doc = docx.Document(str(fpath))
                content = "\n".join(p.text for p in doc.paragraphs)
            except ImportError:
                return f"Word 文件需安装 python-docx 库才能读取：{original_name}"
        else:
            return f"暂不支持的文件类型：{ext}（{original_name}）"

        if not content or not content.strip():
            return f"文件「{original_name}」内容为空。"

        return (
            f"**文件：{original_name}**\n\n"
            f"{content[:3000]}"
            + (f"\n\n*(内容过长，已截断至前 3000 字符)*" if len(content) > 3000 else "")
        )
    except Exception as e:
        return f"读取文件「{original_name}」时出错：{str(e)}"


async def query_experience(query_text: str, category: str = "", top_k: int = 3,
                           user_id: str = "default") -> str:
    """Search historical experiences, optionally filtered by tag/category."""
    results = await experience_service.search_relevant(
        query_text=query_text, user_id=user_id, top_k=top_k, category=category or None,
    )

    if not results:
        return "No relevant historical experiences found."

    lines = [f"**Historical Experiences ({len(results)} found):**\n"]
    for i, r in enumerate(results, 1):
        text = r.get("text", "")
        lines.append(f"[{i}] {text[:500]}")
    return "\n\n".join(lines)


async def use_skill(skill_name: str) -> str:
    """Look up a skill by name and return its body (core workflow) for the agent to follow.

    Returns only the main body (≤100 lines) to keep context lean.
    Hints at the end tell the agent whether detailed references or
    executable scripts are available on demand.
    """
    from app.database import SessionLocal
    from app.models.skill import Skill

    db = SessionLocal()
    try:
        row = db.query(Skill).filter(Skill.title == skill_name).first()
        if not row:
            row = db.query(Skill).filter(Skill.title.ilike(f"%{skill_name}%")).first()
        if not row:
            available = db.query(Skill.title).all()
            names = [r.title for r in available]
            if names:
                return f"未找到技能「{skill_name}」。可用技能：{', '.join(names)}"
            return f"未找到技能「{skill_name}」，且当前没有已配置的技能。"

        parts = [f"**技能：{row.title}**\n"]

        if row.description:
            parts.append(f"*{row.description}*\n")

        parts.append(row.body or row.content)

        # Hint at available on-demand resources
        hints = []
        if row.references:
            hints.append(f'详细参考文档：使用 get_skill_reference("{row.title}") 查阅')
        if row.scripts:
            import json as _json
            try:
                script_list = _json.loads(row.scripts) if isinstance(row.scripts, str) else row.scripts
                names = [s["name"] for s in script_list if isinstance(s, dict) and s.get("name")]
                if names:
                    name_str = "、".join(names)
                    hints.append(
                        f'可执行脚本：{name_str}，使用 '
                        f'run_skill_script("{row.title}", script_name) 执行'
                    )
            except Exception:
                pass

        if hints:
            parts.append("")
            parts.append("---")
            for h in hints:
                parts.append(f"- {h}")

        return "\n".join(parts)
    finally:
        db.close()


async def get_skill_reference(skill_name: str) -> str:
    """Load the detailed reference docs for a skill on demand.

    Skills store their core workflow in 'body' (returned by use_skill)
    and detailed docs in 'references'. Call this only when the user
    asks for more detail or the body says to consult the references.
    """
    from app.database import SessionLocal
    from app.models.skill import Skill

    db = SessionLocal()
    try:
        row = db.query(Skill).filter(Skill.title == skill_name).first()
        if not row:
            row = db.query(Skill).filter(Skill.title.ilike(f"%{skill_name}%")).first()
        if not row:
            return f"未找到技能「{skill_name}」。"
        if not row.references:
            return f"技能「{row.title}」没有参考文档。"
        return (
            f"**参考文档：{row.title}**\n\n"
            f"{row.references}"
        )
    finally:
        db.close()


async def run_skill_script(skill_name: str, script_name: str,
                           content: str = "",
                           timeout: int = 60) -> str:
    """Execute a named script from a skill in the Docker sandbox.

    Looks up the script by name in the skill's scripts JSON, then
    runs it via the sandbox executor. The ``content`` parameter is
    written to ``/sandbox/input.md`` so the script can process it.

    If the script generates output files in ``/sandbox/output/``,
    they are moved to a persistent download directory and a JSON
    result with ``download_id`` is included in the return value.
    """
    from app.database import SessionLocal
    from app.models.skill import Skill
    from app.config import settings
    import json as _json
    import os as _os
    import shutil as _shutil
    from pathlib import Path as _Path
    import uuid as _uuid

    db = SessionLocal()
    try:
        row = db.query(Skill).filter(Skill.title == skill_name).first()
        if not row:
            row = db.query(Skill).filter(Skill.title.ilike(f"%{skill_name}%")).first()
        if not row:
            return f"未找到技能「{skill_name}」。"
        if not row.scripts:
            return f"技能「{row.title}」没有可执行脚本。"

        scripts = _json.loads(row.scripts) if isinstance(row.scripts, str) else row.scripts
        if not isinstance(scripts, list):
            return f"技能「{row.title}」脚本数据格式异常。"

        target = None
        for s in scripts:
            if isinstance(s, dict) and s.get("name") == script_name:
                target = s
                break
        if target is None:
            available = [s.get("name", "?") for s in scripts if isinstance(s, dict)]
            return (
                f"技能「{row.title}」中未找到脚本「{script_name}」。"
                f"可用脚本：{', '.join(available) if available else '无'}"
            )

        code = target.get("code", "")
        if not code or not code.strip():
            return f"脚本「{script_name}」的代码为空。"

        timeout = int(target.get("timeout", timeout))
    finally:
        db.close()

    # Set up extra files and output directory for sandbox
    extra_files: dict[str, str] = {}
    if content:
        extra_files["/sandbox/input.md"] = content

    download_dir = getattr(settings, 'download_dir', str(_Path(__file__).parent.parent.parent / "downloads"))
    _os.makedirs(download_dir, exist_ok=True)
    run_output_dir = _os.path.join(download_dir, f".run-{_uuid.uuid4().hex[:8]}")
    _os.makedirs(run_output_dir, exist_ok=True)
    _os.chmod(run_output_dir, 0o777)

    try:
        from app.sandbox import run_code as _sandbox_run, is_available

        if not is_available():
            return (
                "[Sandbox Unavailable] Docker is not running or sandbox is "
                "disabled on this server. Contact the administrator to enable "
                "the sandbox feature."
            )

        result = await _sandbox_run(
            code,
            timeout=timeout,
            extra_files=extra_files,
            output_dir=run_output_dir,
        )

        # Check if script generated output files
        result_files = []
        for f in _os.listdir(run_output_dir):
            fpath = _os.path.join(run_output_dir, f)
            if _os.path.isfile(fpath):
                result_files.append(f)

        if result_files:
            # Move first output file to persistent download dir
            src = _os.path.join(run_output_dir, result_files[0])
            download_id = _uuid.uuid4().hex[:16]
            ext = _os.path.splitext(result_files[0])[1] or ".docx"
            dest = _os.path.join(download_dir, download_id)
            _shutil.move(src, dest)

            # Write metadata
            meta_path = _os.path.join(download_dir, f"{download_id}.meta")
            _json.dump(
                {"filename": result_files[0]},
                open(meta_path, "w", encoding="utf-8"),
                ensure_ascii=False,
            )

            return _json.dumps(
                {"ok": True, "download_id": download_id, "filename": result_files[0]},
                ensure_ascii=False,
            )

        return result

    finally:
        try:
            _shutil.rmtree(run_output_dir)
        except OSError:
            pass


async def run_code(code: str, timeout: int = 60,
                   extra_files: dict[str, str] | None = None,
                   output_dir: str | None = None) -> str:
    """Execute Python code in a secure Docker sandbox.

    Runs the provided Python code inside an isolated container with
    no network access, limited memory/CPU, and read-only filesystem.
    Use this when a skill workflow requires data analysis, chart
    generation, or computation.

    Args:
        code: Python source code to execute.
        timeout: Max execution time in seconds.
        extra_files: Dict mapping container-path → file-content to mount.
        output_dir: Host directory mounted writable at /sandbox/output.
    """
    from app.sandbox import run_code as _sandbox_run, is_available

    if not is_available():
        return (
            "[Sandbox Unavailable] Docker is not running or sandbox is "
            "disabled on this server. Contact the administrator to enable "
            "the sandbox feature."
        )

    # Defense-in-depth: reject obviously dangerous patterns before
    # spawning a container. Docker isolation is the real guard.
    dangerous = [
        "os.system(", "subprocess.", "shutil.rmtree",
        "__import__(", "compile(", "pty.spawn",
    ]
    code_lower = code.lower()
    for pattern in dangerous:
        if pattern.lower() in code_lower:
            return (
                f"[Sandbox Rejected] Code contains blocked pattern "
                f"'{pattern}'. This is not allowed in the sandbox."
            )

    return await _sandbox_run(code, timeout=timeout,
                               extra_files=extra_files,
                               output_dir=output_dir)


TOOL_FUNCTIONS: dict[str, callable] = {
    "query_rag": query_rag,
    "list_db_connections": list_db_connections,
    "list_db_tables": list_db_tables,
    "query_db": query_db,
    "query_experience": query_experience,
    "find_file_by_name": find_file_by_name,
    "use_skill": use_skill,
    "get_skill_reference": get_skill_reference,
    "run_skill_script": run_skill_script,
    "run_code": run_code,
}

TOOL_SCHEMAS: list[dict] = [
    {
        "type": "function",
        "function": {
            "name": "query_experience",
            "description": "Search historical experiences and knowledge accumulated from past conversations. These are high-quality, user-verified knowledge. Use when the question may have been answered before or involves domain-specific procedures/rules. For general knowledge questions (concepts, principles, common methods), skip this and answer directly.",
            "parameters": {
                "type": "object",
                "properties": {
                    "query_text": {
                        "type": "string",
                        "description": "The question or topic to search for in historical experiences"
                    },
                    "category": {
                        "type": "string",
                        "description": "Leave empty. The system automatically applies the user's selected tag.",
                        "default": ""
                    },
                    "top_k": {
                        "type": "integer",
                        "description": "Number of experiences to return",
                        "default": 3
                    }
                },
                "required": ["query_text"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "find_file_by_name",
            "description": "Search for a specific file by name keyword and return its full content. Use this when the prompt template explicitly instructs you to look for a particular file (e.g., 'if the user mentions turnover, find 离职报告.pdf'). This bypasses semantic search and reads the file directly — use it ONLY when a specific file name is mentioned, NOT for general topic searches.",
            "parameters": {
                "type": "object",
                "properties": {
                    "keyword": {
                        "type": "string",
                        "description": "File name keyword to search for (case-insensitive substring match). E.g. '离职报告' or '考勤制度.pdf'."
                    }
                },
                "required": ["keyword"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "query_rag",
            "description": "Search uploaded documents via the local RAG engine. Use when the question requires looking up information that may exist in the knowledge base (policies, reports, manuals, etc.). For general knowledge questions, skip this and answer from your own training data.",
            "parameters": {
                "type": "object",
                "properties": {
                    "query_text": {
                        "type": "string",
                        "description": "The natural language question or search query"
                    },
                    "category": {
                        "type": "string",
                        "description": "Leave empty. The system automatically applies the user's selected tag. Do NOT guess or invent a category.",
                        "default": ""
                    },
                    "top_k": {
                        "type": "integer",
                        "description": "Number of document chunks to return",
                        "default": 5
                    }
                },
                "required": ["query_text"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "list_db_connections",
            "description": "List all connected database connections (both test and production). Call this FIRST when the user asks about database data, to discover available connections. Returns each connection's ID, name, environment, table name, and field count.",
            "parameters": {
                "type": "object",
                "properties": {},
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "list_db_tables",
            "description": "List all tables and fields for a specific database connection. Call after list_db_connections to explore the schema before writing SQL queries.",
            "parameters": {
                "type": "object",
                "properties": {
                    "connection_id": {
                        "type": "integer",
                        "description": "The connection ID from list_db_connections"
                    }
                },
                "required": ["connection_id"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "use_skill",
            "description": "Load a skill's core workflow (body). Skills are predefined workflows created by the admin. Use this FIRST when the user's request matches a skill's purpose. The returned body contains step-by-step instructions. If the body mentions detailed references or scripts, call get_skill_reference or run_skill_script afterward.",
            "parameters": {
                "type": "object",
                "properties": {
                    "skill_name": {
                        "type": "string",
                        "description": "The name/title of the skill to load."
                    }
                },
                "required": ["skill_name"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_skill_reference",
            "description": "Load detailed reference docs for a skill (on-demand). Only call this when use_skill hints that references are available AND the user needs more detail than the body provides.",
            "parameters": {
                "type": "object",
                "properties": {
                    "skill_name": {
                        "type": "string",
                        "description": "The skill name returned by use_skill."
                    }
                },
                "required": ["skill_name"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "run_skill_script",
            "description": "Execute a named Python script from a skill in the Docker sandbox. Call this when a skill workflow instructs you to run a specific script. The sandbox has pandas, numpy, matplotlib, plotly, openpyxl, tabulate, python-docx pre-installed. Pass the content to process via the 'content' parameter — it will be written to /sandbox/input.md for the script to read.",
            "parameters": {
                "type": "object",
                "properties": {
                    "skill_name": {
                        "type": "string",
                        "description": "The skill name returned by use_skill."
                    },
                    "script_name": {
                        "type": "string",
                        "description": "The script file name to execute (e.g. 'format_official_docx.py')."
                    },
                    "content": {
                        "type": "string",
                        "description": "The content (e.g. markdown text) to pass to the script. Written to /sandbox/input.md.",
                        "default": ""
                    },
                    "timeout": {
                        "type": "integer",
                        "description": "Maximum execution time in seconds (default 60).",
                        "default": 60
                    }
                },
                "required": ["skill_name", "script_name"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "query_db",
            "description": "Execute a read-only SQL query (SELECT, SHOW, DESCRIBE, EXPLAIN) on a specific database connection. Call list_db_connections first to get the connection ID.",
            "parameters": {
                "type": "object",
                "properties": {
                    "sql_query": {
                        "type": "string",
                        "description": "A read-only SQL SELECT/SHOW/DESCRIBE/EXPLAIN query"
                    },
                    "connection_id": {
                        "type": "integer",
                        "description": "The connection ID from list_db_connections"
                    }
                },
                "required": ["sql_query", "connection_id"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "run_code",
            "description": "Execute Python code in a secure Docker sandbox with no network access. Use this when a skill workflow requires data analysis, chart generation, or computation. The sandbox has pandas, numpy, matplotlib, plotly, openpyxl, and tabulate pre-installed. Use print() to output results.",
            "parameters": {
                "type": "object",
                "properties": {
                    "code": {
                        "type": "string",
                        "description": "The Python 3.12 source code to execute in the sandbox. Use print() to output results. Avoid os.system, subprocess, or file I/O — the filesystem is read-only and network is disabled."
                    },
                    "timeout": {
                        "type": "integer",
                        "description": "Maximum execution time in seconds (default 60, max 120).",
                        "default": 60
                    }
                },
                "required": ["code"]
            }
        }
    },
]


def get_schemas() -> list[dict]:
    """Get tool schemas for OpenAI tool-calling API."""
    return TOOL_SCHEMAS


async def execute_tool(name: str, user_id: str = "default",
                       kb_scope: str = "none",
                       db_scope: list[int] | None = None,
                       default_category: str = "",
                       **kwargs) -> str:
    """Execute a tool function by name with given arguments.

    kb_scope / db_scope are forwarded to the individual tools so they can
    enforce the permissions configured by the admin in user management.

    default_category is used for query_rag when the LLM doesn't specify a
    category — typically set from the prompt template the user selected.
    """
    func = TOOL_FUNCTIONS.get(name)
    if func is None:
        return f"[Unknown tool: {name}]"

    try:
        if name == "query_rag":
            kwargs["user_id"] = user_id
            kwargs["kb_scope"] = kb_scope
            # User's selected tag always overrides AI-invented category
            if default_category:
                kwargs["category"] = default_category
        elif name == "query_experience":
            kwargs["user_id"] = user_id
            if default_category:
                kwargs["category"] = default_category
        elif name == "find_file_by_name":
            kwargs["user_id"] = user_id
        elif name in ("list_db_connections", "list_db_tables", "query_db"):
            kwargs["db_scope"] = db_scope
        result = await func(**kwargs)
        return result
    except Exception as e:
        return f"[Tool Error: {name}] {str(e)}"
